"""
    @file:   translate_excel.py
    @brief:  该脚本通过调用大模型接口，将excel中使用的语言翻译为指定语言，翻译后输出到output.xlsx 
"""


import openai
from dotenv import load_dotenv, find_dotenv
import os
import openpyxl

# 加载.env中的环境变量
load_dotenv(find_dotenv())

# 从环境变量加载必要参数
openai.api_key = os.getenv("OPENAI_API_KEY")
openai.api_base = os.getenv("OPENAI_API_BASE")
openai_model = os.getenv("OPENAI_MODEL")
# 输入文件
input_file = os.getenv("CONFIG_FILE_INPUT")
# 输出文件
output_file = os.getenv("CONFIG_FILE_OUTPUT")
# 需要翻译成的目标语言
target_lang = os.getenv("CONFIG_TARGET_LANG")
# 需要排除的字段
expected_for_word = ["admin"]

def create_system_prompt(targetLang):
    """
    创建系统提示信息，用于指导翻译模型
    :param targetLang: 目标语言
    :return: 系统提示字符串
    """
    return f"你现在是熟练的翻译大师，请将我<输入的文本>翻译为<"+targetLang+">,你只需要提供翻译的结果即可。请严格遵循以下翻译要求：\n"\
            "1. 当<输入的文本>已经是目标语言<"+targetLang+">,请发送<输入的文本>\n"\
            "2. 当<输入的文本>是"+",".join(expected_for_word)+",其中任意一个的时候,请发送<输入的文本>\n"\
            "3. 除了翻译<输入的文本>为目标语言<"+targetLang+">并发送给我之外，不要携带任何和翻译文本无关信息。"

def translateTo(text, targetLang, model) -> str:
    """
    使用OpenAI API进行文本翻译
    :param text: 需要翻译的文本
    :param targetLang: 目标语言
    :param model: OpenAI模型名称
    :return: 翻译后的文本
    """
    response = openai.ChatCompletion.create(
        model=model,
        messages=[
            {"role": "system", "content": create_system_prompt(targetLang)},
            {"role": "user", "content": text}
        ],
        # 禁用流式传输
        stream=False, 
        temperature=1.1
    )
    return response["choices"][0]["message"]["content"]

if __name__ == "__main__":
    # 加载输入的Excel文件
    wb = openpyxl.load_workbook(input_file)
    # 记录总共需要翻译的单元格
    total_cells = 0
    # 记录已经翻译的单元格
    translated_cells = 0
    # 每翻译batch_size格单元格保存一次文件
    batch_size = 5

    # 计算总共要翻译的单元格个数
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                # 确保当前单元格是一个字符串
                if isinstance(cell.value, str):
                    total_cells += 1

    try:
        for sheet in wb:
            # 遍历所有行和列
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        # 确保当前单元格是一个字符串
                        if isinstance(cell.value, str):
                            cell_old = cell.value
                            cell.value = translateTo(cell.value, target_lang, openai_model)
                            translated_cells += 1
                            if cell_old != cell.value:
                                print(f"from \"{cell_old}\" to \"{cell.value}\" processed ({translated_cells}/{total_cells})")
                                # 每翻译batch_size个单元格后保存一次文件
                                if translated_cells % batch_size == 0 and translated_cells != 0:
                                    wb.save(output_file)
        # 翻译完成后保存文件
        wb.save(output_file)
    except Exception as e:
        print(f"出现错误！{e}")
        wb.save(output_file)

    print("翻译完成！")